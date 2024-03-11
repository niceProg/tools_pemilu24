from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import Font
from selenium.common.exceptions import TimeoutException, WebDriverException
from fake_useragent import UserAgent
import time

file_excel_path = "..."  # Ganti dengan nama file Excel

def initialize_browser():
    ua = UserAgent()
    chrome_options = Options()
    chrome_options.add_argument(f'user-agent={ua.random}')
    chrome_options.add_argument("--no-sandbox")  # Tambahkan opsi no-sandbox untuk stabilitas
    chrome_options.add_argument("--disable-dev-shm-usage")  # Mengurangi penggunaan resource
    try:
        driver = webdriver.Chrome(options=chrome_options)
    except WebDriverException as e:
        print(f"Gagal menginisialisasi WebDriver: {e}")
        exit(1)
    return driver

def close_and_reopen_browser(driver):
    try:
        driver.quit()
    except Exception as e:
        print(f"Error saat menutup browser: {e}")
    time.sleep(3)  # Beri jeda untuk memastikan browser benar-benar tertutup
    return initialize_browser()

driver = initialize_browser()
time.sleep(3)

def convert_nik_to_str(nik):
    return str(nik)

def is_valid_nik(nik):
    # Contoh validasi sederhana, perlu disesuaikan
    return len(nik) == 16

def get_text_from_sibling(driver, xpath):
    return driver.execute_script(
        f"return document.evaluate(\"{xpath}\", document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue.textContent"
    )

def wait_for_element(driver, by, value, timeout=5):
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_all_elements_located((by, value))
        )
    except TimeoutException:
        return False
    return True

def slow_type(element, text, delay=0.2):
    """Fungsi untuk mengetik teks pada elemen secara perlahan."""
    for character in text:
        element.send_keys(character)
        time.sleep(delay)

def cek_data(driver, nik_list):
    result_data = []

    for nik in nik_list:
        if not is_valid_nik(nik):
            result_data.append((nik, "NIK Tidak Valid", "", "", "", ""))
            continue

        data_found = False  # Flag untuk menandakan apakah data sudah ditemukan
        captcha_encountered = False  # Flag untuk menandakan apakah captcha ditemukan

        for attempt in range(1):  # Max retries
            while True:  # Loop untuk mengatasi captcha
                try:
                    driver.get("https://cekdptonline.kpu.go.id/")
                    input_nik = driver.find_element(By.XPATH, "//input[@class='form-control is-valid']")
                    slow_type(input_nik, nik)
                    search_button = driver.find_element(By.XPATH, "//button[./span[contains(text(), 'Pencarian')]]")
                    search_button.click()

                    # Cek untuk invalid captcha
                    invalid_captcha = get_text_from_sibling(driver, "//h2[@class='mb-2']/b")
                    if invalid_captcha == "Oops, Something went wrong..":
                        print("Invalid captcha encountered. Retrying...")
                        captcha_encountered = True
                        continue  # Skip ke iterasi selanjutnya dari while loop untuk mengatasi captcha

                    if wait_for_element(driver, By.XPATH, "//div[@class='column']"):
                        # Elemen yang diperlukan berhasil dimuat, keluar dari loop captcha
                        captcha_encountered = False
                        break
                except Exception as e:
                    print(f"Error: {e}")
                    break  # Keluar dari loop jika terjadi kesalahan

            if captcha_encountered:
                # Jika loop dihentikan karena captcha, lanjutkan ke percobaan selanjutnya
                continue

            # Logika pengambilan data jika captcha berhasil diatasi
            if wait_for_element(driver, By.XPATH, "//div[@class='column']"):
                nama = get_text_from_sibling(driver, "//div[@class='column']//div[@class='row row-1']/p/span[contains(text(), 'Nama Pemilih')]/following-sibling::text()")
                tps = get_text_from_sibling(driver, "//div[@class='column']//div[@class='row row-1']/p/span[contains(text(), 'TPS')]/following-sibling::text()")
                keterangan = get_text_from_sibling(driver, "//div[@class='column']//div[@class='row row-3']/p[@class='row--right']/span[contains(text(), 'Kelurahan')]/following-sibling::text()")
                kabupaten = get_text_from_sibling(driver, "//div[@class='column']//div[@class='row row-3']/p[@class='row--left']/span[contains(text(), 'Kabupaten')]/following-sibling::text()")
                alamat = get_text_from_sibling(driver, "//div[@class='column']//div[@class='row row-3']/p[@class='row--left']/span[contains(text(), 'Alamat Potensial TPS')]/following-sibling::span")

                if all(val.strip() for val in [nama, tps, keterangan, kabupaten, alamat]):
                    result_data.append((nik, nama, tps, keterangan, kabupaten, alamat))
                    print(f"Data untuk NIK {nik} berhasil diambil. Menunggu 10 detik sebelum melanjutkan...")
                    time.sleep(5)  # Delay antara permintaan untuk menghindari rate limit
                    data_found = True  # Set flag bahwa data ditemukan
                    break  # Keluar dari loop karena data valid ditemukan

        if not data_found:
            # Jika setelah max retries (dan upaya mengatasi captcha) data tidak ditemukan
            result_data.append((nik, "Data Tidak Ditemukan", "", "", "", ""))

    return result_data

def process_sheet(sheet, driver):
    nik_list = [convert_nik_to_str(sheet[f'B{row_index}'].value) for row_index in range(6, sheet.max_row + 1)]
    total_nik = len(nik_list)
    nik_processed = 0

    nik_to_row = {nik: row_index for row_index, nik in enumerate(nik_list, start=6)}

    batch_size = 20  # Sesuaikan batch_size sesuai dengan kebutuhan
    for i in range(0, len(nik_list), batch_size):
        batch_nik_list = nik_list[i:i+batch_size]
        hasil_cek = cek_data(driver, batch_nik_list)

        for nik, nama, tps, keterangan, kabupaten, alamat in hasil_cek:
            row_index = nik_to_row.get(nik)
            if row_index:
                sheet[f'D{row_index}'] = nama
                sheet[f'E{row_index}'] = tps
                sheet[f'F{row_index}'] = keterangan
                sheet[f'G{row_index}'] = kabupaten
                sheet[f'H{row_index}'] = alamat

                if nama in ["Data Tidak Ditemukan", "NIK Tidak Valid"]:
                    sheet[f'B{row_index}'].font = Font(color="FF0000")
            nik_processed += 1
            print(f"Progress: {nik_processed} dari {total_nik} NIK diproses ({(nik_processed/total_nik)*100:.2f}%)")
        
        driver = close_and_reopen_browser(driver)  # Tutup dan buka kembali browser setiap batch

workbook = load_workbook(file_excel_path)
sheets_to_process = ["..."] #tentukan nama sheet yang ada di excel

for sheet_name in workbook.sheetnames:
    if sheet_name in sheets_to_process:
        print(f"Processing sheet: {sheet_name}")
        process_sheet(workbook[sheet_name], driver)

driver.quit()
workbook.save(file_excel_path)