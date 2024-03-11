from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from collections import Counter

def count_classifications(sheet):
    nik_count = Counter()
    classifications = {'valid': 0, 'double': 0, 'invalid': 0, 'non_jaksel': 0}

    # Pertama, hitung semua NIK dan tandai apakah mereka invalid atau non-jaksel
    nik_details = {}
    for row in sheet.iter_rows(min_row=7, values_only=True):
        nik = row[1]
        nama = row[3]
        kabupaten = row[6]

        nik_count[nik] += 1
        if nama in ["NIK Tidak Valid", "Data Tidak Ditemukan"]:
            nik_details[nik] = 'invalid'
        elif kabupaten != 'KOTA JAKARTA SELATAN':
            nik_details[nik] = 'non_jaksel'

    # Kemudian, klasifikasikan setiap NIK berdasarkan detail dan jumlah kemunculannya
    for nik, count in nik_count.items():
        detail = nik_details.get(nik)

        if detail == 'invalid':
            classifications['invalid'] += 1
        elif detail == 'non_jaksel':
            classifications['non_jaksel'] += 1
        else:
            if count > 1:
                # NIK pertama dianggap valid, sisanya dianggap double
                classifications['valid'] += 1
                classifications['double'] += (count - 1)
            else:
                classifications['valid'] += 1

    return classifications

def create_bar_chart(workbook, classifications):
    sheet = workbook.create_sheet(title="TOTAL")
    rows = [
        ['Klasifikasi', 'Jumlah'],
        ['NIK Valid', classifications['valid']],
        ['NIK Double', classifications['double']],
        ['NIK Invalid', classifications['invalid']],
        ['NIK Non Jakarta Selatan', classifications['non_jaksel']],
    ]
    for row in rows:
        sheet.append(row)

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Jumlah Klasifikasi NIK"
    chart.y_axis.title = 'Jumlah'
    chart.x_axis.title = 'Klasifikasi'

    data = Reference(sheet, min_col=2, min_row=1, max_row=5, max_col=2)
    cats = Reference(sheet, min_col=1, min_row=2, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    sheet.add_chart(chart, "D2")

file_path = "..." #file input
workbook = load_workbook(file_path)
main_sheet = workbook.active  # Atau ganti dengan nama sheet yang diinginkan

classifications = count_classifications(main_sheet)
create_bar_chart(workbook, classifications)

workbook.save("...") #file output
