import os
from openpyxl import load_workbook, Workbook

def get_data_from_sheet(sheet, include_header=False):
    """
    Mengambil semua data dari sheet.
    Jika include_header=True, mengambil header dari sheet pertama saja.
    """
    if include_header:
        return list(sheet.iter_rows(values_only=True))
    else:
        return list(sheet.iter_rows(values_only=True, min_row=6))

def combine_sheets(input_folder_path, output_file_path):
    """
    Menggabungkan sheet dari beberapa file Excel (.xlsx) dalam satu folder ke dalam satu sheet.
    Mengabaikan 5 baris pertama dari setiap sheet kecuali yang pertama.
    """
    output_wb = Workbook()
    output_sheet = output_wb.active
    output_sheet.title = "Combined Data"

    first_file = True  # Flag untuk menandai file pertama

    for filename in os.listdir(input_folder_path):
        if filename.endswith(".xlsx"):
            input_file_path = os.path.join(input_folder_path, filename)
            input_wb = load_workbook(input_file_path, read_only=True)

            for sheet_name in input_wb.sheetnames:
                sheet = input_wb[sheet_name]
                data = get_data_from_sheet(sheet, include_header=first_file)

                for row in data:
                    # Tambahkan baris langsung tanpa menambahkan nomor urut
                    output_sheet.append(row)

                first_file = False  # Setelah file pertama, ubah flag menjadi False

    output_wb.save(output_file_path)
    print(f"Data combined and saved to {output_file_path}")

# Set the input folder and output file paths here
input_folder_path = ""  # Ganti dengan path folder Excel input Anda
output_file_path = ""  # Ganti dengan path file Excel output Anda

combine_sheets(input_folder_path, output_file_path)