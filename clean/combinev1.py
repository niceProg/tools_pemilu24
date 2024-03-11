from openpyxl import load_workbook, Workbook
import tkinter as tk
from tkinter import filedialog

def get_file_paths_gui():
    root = tk.Tk()
    root.withdraw()  # Menyembunyikan jendela tkinter utama
    print("Pilih file Excel:")
    file_paths = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])  # Meminta pengguna untuk memilih beberapa file Excel
    print("Simpan file output sebagai (pastikan menambahkan .xlsx pada akhir nama file):")
    output_path = filedialog.asksaveasfilename(defaultextension=".xlsx")  # Meminta pengguna untuk memilih lokasi penyimpanan file output
    return file_paths, output_path  # Mengembalikan daftar path dan path output sebagai string

def combine_sheets(input_file_paths, output_file_path):
    output_wb = Workbook()
    output_sheet = output_wb.active
    output_sheet.title = "Combined Data"
    current_number = 1  # Inisialisasi nomor urut
    first_sheet = True

    for input_file_path in input_file_paths:  # Iterasi melalui setiap file input
        input_wb = load_workbook(input_file_path)
        for sheet_name in input_wb.sheetnames:
            sheet = input_wb[sheet_name]
            data = list(sheet.iter_rows(values_only=True))

            for row_index, row in enumerate(data, start=1):
                if first_sheet:
                    if row_index < 6:
                        # Untuk baris ke-1 sampai ke-5 di sheet pertama, tambahkan tanpa nomor urut
                        output_sheet.append(row)
                    else:
                        # Mulai memberi nomor urut dari baris ke-6 di sheet pertama
                        output_sheet.append([current_number] + list(row[1:]))
                        current_number += 1
                else:
                    if row_index > 5:
                        # Untuk sheet kedua dan seterusnya, tambahkan baris dengan nomor urut
                        output_sheet.append([current_number] + list(row[1:]))
                        current_number += 1

            first_sheet = False  # Setelah melewati sheet pertama, ubah menjadi False

    output_wb.save(output_file_path)
    print(f"Data combined and saved to {output_file_path}")

input_file_paths, output_file_path = get_file_paths_gui()
combine_sheets(input_file_paths, output_file_path)
