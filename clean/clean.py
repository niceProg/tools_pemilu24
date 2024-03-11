from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from collections import Counter
import tkinter as tk
from tkinter import filedialog

def write_row(sheet, row_number, data, font, fill=None, tps_format=False):
    for j, value in enumerate(data, start=2):  # Mulai dari kolom B
        cell = sheet.cell(row=row_number, column=j)
        cell.value = value if value != float('inf') else "N/A"
        cell.font = font
        if fill:
            cell.fill = fill
        if j == 5 and tps_format:  # Jika kolom E dan format TPS diperlukan
            cell.number_format = '0'

def process_data(sheet, data, start_row, font, nik_count, nik_processed, outside_color=None, duplicate_color='FFFF00'):
    for data_tuple in data:
        nik = data_tuple[0]
        # Inisialisasi fill sebagai None
        fill = None

        # Cek apakah NIK ini duplikat dan bukan kemunculan pertama
        is_double = nik_count[nik] > 1 and nik_processed[nik] > 0
        if is_double:
            # Jika NIK duplikat dan bukan yang pertama, gunakan warna duplikat
            fill = PatternFill(start_color=duplicate_color, end_color=duplicate_color, fill_type='solid')
        
        # Jika outside_color didefinisikan dan ini adalah kemunculan pertama dari NIK duplikat atau bukan NIK duplikat
        if outside_color and (nik_processed[nik] == 0 or not is_double):
            # Gunakan warna luar jika disediakan
            fill = PatternFill(start_color=outside_color, end_color=outside_color, fill_type='solid')

        write_row(sheet, start_row, data_tuple, font, fill, tps_format=True)
        nik_processed[nik] += 1
        start_row += 1

def sort_data(data_list, key_index):
        return sorted(data_list, key=lambda x: (x[key_index] if x[key_index] is not None else "", x[0] if x[0] is not None else ""))

def sort_and_color_excel(sheet):
    valid_data, non_jakarta, invalid_data = [], [], []
    nik_count = Counter()
    nik_processed = {}

    for row in sheet.iter_rows(min_row=7, values_only=True):
        nik, no_hp, nama, tps, kelurahan, kabupaten, alamat = row[1:8]
        nik_count[nik] += 1
        nik_processed[nik] = 0  # Initialize or reset the counter for each NIK
        try:
            tps = int(tps) if tps is not None else float('inf')
        except (ValueError, TypeError):
            tps = float('inf')
        data_tuple = (nik, no_hp, nama, tps, kelurahan, kabupaten, alamat)
        if nama in ["NIK Tidak Valid", "Data Tidak Ditemukan"]:
            invalid_data.append(data_tuple)
        elif kabupaten != 'KOTA JAKARTA SELATAN':
            non_jakarta.append(data_tuple)
        else:
            valid_data.append(data_tuple)

    black_font = Font(color='000000')
    abu = 'acb5af'
    current_row = 6
    # Proses dan tulis data valid
    process_data(sheet, sort_data(valid_data, 3), current_row, black_font, nik_count, nik_processed)
    current_row += len(valid_data)
    # Proses dan tulis data non-Jakarta
    process_data(sheet, sort_data(non_jakarta, 3), current_row, black_font, nik_count, nik_processed, outside_color=abu)
    current_row += len(non_jakarta)

    # Proses dan tulis data invalid
    for data_tuple in invalid_data:
        write_row(sheet, current_row, data_tuple, black_font, PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'), tps_format=False)
        current_row += 1

def get_file_path_gui():
    root = tk.Tk()
    root.withdraw()  # Menyembunyikan jendela tkinter utama

    print("Pilih file Excel:")
    file_path = filedialog.askopenfilename()  # Meminta pengguna untuk memilih file Excel
    print("Simpan file output sebagai (pastikan menambahkan .xlsx pada akhir nama file):")
    output_path = filedialog.asksaveasfilename(defaultextension=".xlsx")  # Meminta pengguna untuk memilih lokasi penyimpanan file output

    return file_path, output_path

# Mendapatkan path file input dan output dari pengguna
input_file_path, output_file_path = get_file_path_gui()

workbook = load_workbook(input_file_path)
sheet_name = 'Combined Data'  # Ganti dengan nama sheet yang diinginkan

if sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    sort_and_color_excel(sheet)
else:
    print(f"Sheet {sheet_name} tidak ditemukan di workbook.")

workbook.save(output_file_path)