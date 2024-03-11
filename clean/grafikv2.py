from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference
from collections import Counter
import tkinter as tk
from tkinter import filedialog

def get_file_paths_gui():
    root = tk.Tk()
    root.withdraw()  # Menyembunyikan jendela tkinter utama

    print("Pilih file Excel:")
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")], title="Pilih file Excel")
    print("Simpan file output sebagai (pastikan menambahkan .xlsx pada akhir nama file):")
    output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", title="Simpan file output sebagai")

    return file_path, output_path

def count_votes_by_region(sheet, kelurahan_terpilih):
    vote_count = Counter()

    for row in sheet.iter_rows(min_row=7, values_only=True):
        nik, nama, kelurahan = row[1], row[3], row[5]  # Sesuaikan indeks kolom dengan data Anda
        # Pastikan kelurahan adalah string sebelum menerapkan .strip() dan .upper()
        kelurahan = kelurahan.strip().upper() if kelurahan else ""  # Normalisasi nama kelurahan

        if nama not in ["NIK Tidak Valid", "Data Tidak Ditemukan"] and kelurahan in kelurahan_terpilih:
            vote_count[kelurahan] += 1

    return vote_count

def create_bar_chart(workbook, vote_count):
    if not vote_count:
        print("Tidak ada data valid untuk ditampilkan.")
        return

    sheet = workbook.create_sheet(title="Vote Summary")
    rows = [['Wilayah', 'Jumlah Suara']] + [[region, count] for region, count in vote_count.items()]
    for row in rows:
        sheet.append(row)

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Jumlah Suara per Kelurahan Terpilih"
    chart.y_axis.title = 'Jumlah Suara'
    chart.x_axis.title = 'Kelurahan'

    data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row, max_col=2)
    cats = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    sheet.add_chart(chart, "D2")

input_file_path, output_file_path = get_file_paths_gui()

if input_file_path and output_file_path:
    workbook = load_workbook(input_file_path)
    main_sheet = workbook.active  # Atau ganti dengan nama sheet yang diinginkan

    # Tentukan kelurahan yang ingin ditampilkan
    kelurahan_terpilih = ["Cilandak Barat", "Cipete Selatan", "Gandaria Selatan", "Lebak Bulus", "Pondok Labu",
                          "Cipete Utara", "Gandaria Utara", "Gunung", "Kramat Pela", "Melawai", "Petogogan", "Pulo", "Rawa Barat", "Selong", "Senayan",
                          "Grogol Selatan", "Grogol Utara", "Kebayoran Lama Selatan", "Kebayoran Lama Utara", "Cipulir", "Pondok Pinang",
                          "Bintaro", "Pesanggrahan", "Petukangan Selatan", "Petukangan Utara", "Ulujami",
                          "Guntur", "Karet Kuningan", "Karet Semanggi", "Karet", "Kuningan Timur", "Menteng Atas", "Pasar Manggis", "Setiabudi"
                          ]
    kelurahan_terpilih = [k.upper() for k in kelurahan_terpilih]  # Pastikan daftar kelurahan dalam uppercase untuk konsistensi

    vote_count = count_votes_by_region(main_sheet, kelurahan_terpilih)
    create_bar_chart(workbook, vote_count)

    workbook.save(output_file_path)
    print(f"Data grafik untuk kelurahan terpilih telah disimpan ke {output_file_path}")
else:
    print("Operasi dibatalkan.")